from openpyxl import load_workbook

# Ruta del formato original y vacío del informe de resultados
ruta_formato_original = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\03_Muestreo_Superficial\02_Visual_SC\01_Input\DPT-F23 Informe de resultados V03 - propuesta.xlsx"
# Ruta donde se guarda el informe modificado y diligenciado
ruta_formato_modificado = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\03_Muestreo_Superficial\02_Visual_SC\02_Output"
# Ruta del archivo que contiene los datos para el diccionario
ruta_archivo_fuente_1 = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\03_Muestreo_Superficial\02_Visual_SC\01_Input\DPT-F11MuestreoDeAguaSuperficial-034.xlsx"
# Ruta del archivo que contiene los datos para la primera parte del encabezado (Solicitud)
ruta_archivo_fuente_2 = r"C:\Users\SHI-PC34.SHI-PC34\Desktop\Martin_Anaya\03_Muestreo_Superficial\02_Visual_SC\01_Input\Solicitud_Toma_Muestra_Parametros_Campo_034_2023.xlsx"

# Cargar el formato de Excel original
libro_original = load_workbook(ruta_formato_original)
# Leer archivo fuente 1
libro_fuente_1 = load_workbook(ruta_archivo_fuente_1)
# Leer archivo fuente 2
libro_fuente_2 = load_workbook(ruta_archivo_fuente_2)

# Modificar archivo fuente
hoja_fuente_1 = libro_fuente_1.active
# Modificar archivo fuente
hoja_fuente_2 = libro_fuente_2.active 

#Contador para el número de filas llenas
filas_llenas = 0
#Iteración sobre las y conteo de las filas llenas
for fila in hoja_fuente_1.iter_rows():
    if not all(cell is None for cell in fila):
        filas_llenas += 1
print()
print("Número de puntos de muestreo:", filas_llenas - 1)
print("Desde el número 2 hasta el número", filas_llenas)
print()

datos_encabezado = {"Proyecto":["D9","I7",""],"Centro_costos":["D10","I9",""],"Numero_solicitud":["G4","H3",""],"Plan_muestreo":["G4","I10",""],"Cliente":["D7","C7",""],"NIT_CC":["D8","C9",""],"Direccion":["D11","C10",""],"Contacto":["D12","C11",""],"Telefono":["D13","C13",""],"Correo":["D14","C14",""],"Responsable_1":["K2","I11",""],"Responsable_2":["L2","I12",""]}
datos_parametros = {"Departamento":["H","A",""],"Municipio":["I","B",""],"Fecha":["M","F",""],"Hora":["N","G",""],"Codigo_punto":["O","E",""],"Nombre_fuente":["R","C",""]}
datos_parametros_2 = {"pH":["AL","K",""],"Temperatura":["AM","K",""],"OD":["AP","K",""],"Conductividad_E":["AV","K",""],"Caudal":["BG","K",""]}

nombre = "\Informe_llenado"

# Crear copia del archivo original
libro_modificado = load_workbook(ruta_formato_original)
# Modificar copia del archivo original
# Modificar el archivo copiado
nombre_hoja_modificada = 'Parte 1'
hoja_modificada = libro_modificado.get_sheet_by_name(nombre_hoja_modificada)
hoja_modificada = libro_modificado.active 

lista_keys = list(datos_encabezado.keys())
lista_keys_2 = list(datos_parametros.keys())
lista_keys_3 = list(datos_parametros_2.keys())              #Extraer una lista con las "claves" del diccionario

for dato in range(len(datos_encabezado)):
    
    key = lista_keys[dato]                              #Definir cada una de las claves, posición dato de la lista lista_keys
    celda = datos_encabezado[key][0]                    #Obtener la posición de la celda y cocnatenar con el número de la fila
    valor = hoja_fuente_2[celda].value

    if valor is not None:
        datos_encabezado[key][2] = valor
    else:                
        valor = hoja_fuente_1[celda].value
        datos_encabezado[key][2] = valor

    if key == "Numero_solicitud":
        datos_encabezado[key][2] = valor+" - 1"
        
    hoja_modificada[datos_encabezado[key][1]] = datos_encabezado[key][2]

a = 19
b = 19

for fila in range(2,7):
    
    for dato in range(len(datos_parametros)):

        key = lista_keys_2[dato]
        celda_2 = datos_parametros[key][0]+str(fila)
        valor = hoja_fuente_1[celda_2].value
        datos_parametros[key][2] = valor
        hoja_modificada[datos_parametros[key][1] + str(a)] = valor

    for dato in range(len(datos_parametros_2)):
            
        key = lista_keys_3[dato]
        celda_3 = datos_parametros_2[key][0]+str(fila)
        valor = hoja_fuente_1[celda_3].value
        datos_parametros_2[key][2] = valor
        hoja_modificada[datos_parametros_2[key][1] + str(b)] = valor

        b = b + 1
    a = a + 5

ruta_ultima = ruta_formato_modificado+nombre+".xlsx"            
libro_modificado.save(ruta_ultima)
libro_modificado.close()

##############################################################---PARA LA SEGUNDA HOJA---#######################################################################
if 6 < filas_llenas:
    
    libro_modificado_2 = load_workbook(ruta_ultima)
    nombre_hoja_modificada_2 = 'Parte 2'
    hoja_modificada_2 = libro_modificado_2.get_sheet_by_name(nombre_hoja_modificada_2)

    for dato in range(len(datos_encabezado)):
        
        key = lista_keys[dato]                              #Definir cada una de las claves, posición dato de la lista lista_keys
        celda = datos_encabezado[key][0]                    #Obtener la posición de la celda y cocnatenar con el número de la fila
        valor = hoja_fuente_2[celda].value

        if valor is not None:
            datos_encabezado[key][2] = valor
        else:                
            valor = hoja_fuente_1[celda].value
            datos_encabezado[key][2] = valor

        if key == "Numero_solicitud":
            datos_encabezado[key][2] = valor+" - 2"

        hoja_modificada_2[datos_encabezado[key][1]] = datos_encabezado[key][2]

    a = 19
    b = 19
    
    for fila in range(7,12):

        for dato in range(len(datos_parametros)):
        
            key = lista_keys_2[dato]
            celda_2 = datos_parametros[key][0]+str(fila)
            valor = hoja_fuente_1[celda_2].value
            datos_parametros[key][2] = valor
            hoja_modificada_2[datos_parametros[key][1] + str(a)] = valor

        for dato in range(len(datos_parametros_2)):
        
            key = lista_keys_3[dato]
            celda_3 = datos_parametros_2[key][0]+str(fila)
            valor = hoja_fuente_1[celda_3].value
            datos_parametros_2[key][2] = valor
            hoja_modificada_2[datos_parametros_2[key][1] + str(b)] = valor

            b = b + 1
        a = a + 5

    ruta_ultima_2 = ruta_formato_modificado+nombre+".xlsx"            
    libro_modificado_2.save(ruta_ultima_2)
    libro_modificado_2.close()    

##############################################################---PARA LA TERCERA HOJA---#######################################################################
if 11 < filas_llenas:
    
    libro_modificado_3 = load_workbook(ruta_ultima)
    nombre_hoja_modificada_3 = 'Parte 3'
    hoja_modificada_3 = libro_modificado_3.get_sheet_by_name(nombre_hoja_modificada_3)

    for dato in range(len(datos_encabezado)):
        
        key = lista_keys[dato]                              #Definir cada una de las claves, posición dato de la lista lista_keys
        celda = datos_encabezado[key][0]                    #Obtener la posición de la celda y cocnatenar con el número de la fila
        valor = hoja_fuente_2[celda].value

        if valor is not None:
            datos_encabezado[key][2] = valor
        else:                
            valor = hoja_fuente_1[celda].value
            datos_encabezado[key][2] = valor

        if key == "Numero_solicitud":
            datos_encabezado[key][2] = valor+" - 3"

        hoja_modificada_3[datos_encabezado[key][1]] = datos_encabezado[key][2]

    a = 19
    b = 19
    
    for fila in range(12,17):

        for dato in range(len(datos_parametros)):
        
            key = lista_keys_2[dato]
            celda_2 = datos_parametros[key][0]+str(fila)
            valor = hoja_fuente_1[celda_2].value
            datos_parametros[key][2] = valor
            hoja_modificada_3[datos_parametros[key][1] + str(a)] = valor

        for dato in range(len(datos_parametros_2)):
        
            key = lista_keys_3[dato]
            celda_3 = datos_parametros_2[key][0]+str(fila)
            valor = hoja_fuente_1[celda_3].value
            datos_parametros_2[key][2] = valor
            hoja_modificada_3[datos_parametros_2[key][1] + str(b)] = valor

            b = b + 1
        a = a + 5

    ruta_ultima_3 = ruta_formato_modificado+nombre+".xlsx"            
    libro_modificado_3.save(ruta_ultima_3)
    libro_modificado_3.close() 

##############################################################---PARA LA CUARTA HOJA---#######################################################################
if 16 < filas_llenas:
    
    libro_modificado_4 = load_workbook(ruta_ultima)
    nombre_hoja_modificada_4 = 'Parte 4'
    hoja_modificada_4 = libro_modificado_4.get_sheet_by_name(nombre_hoja_modificada_4)

    for dato in range(len(datos_encabezado)):
        
        key = lista_keys[dato]                              #Definir cada una de las claves, posición dato de la lista lista_keys
        celda = datos_encabezado[key][0]                    #Obtener la posición de la celda y cocnatenar con el número de la fila
        valor = hoja_fuente_2[celda].value

        if valor is not None:
            datos_encabezado[key][2] = valor
        else:                
            valor = hoja_fuente_1[celda].value
            datos_encabezado[key][2] = valor

        if key == "Numero_solicitud":
            datos_encabezado[key][2] = valor+" - 4"

        hoja_modificada_4[datos_encabezado[key][1]] = datos_encabezado[key][2]

    a = 19
    b = 19
    
    for fila in range(17,22):

        for dato in range(len(datos_parametros)):
        
            key = lista_keys_2[dato]
            celda_2 = datos_parametros[key][0]+str(fila)
            valor = hoja_fuente_1[celda_2].value
            datos_parametros[key][2] = valor
            hoja_modificada_4[datos_parametros[key][1] + str(a)] = valor

        for dato in range(len(datos_parametros_2)):
        
            key = lista_keys_3[dato]
            celda_3 = datos_parametros_2[key][0]+str(fila)
            valor = hoja_fuente_1[celda_3].value
            datos_parametros_2[key][2] = valor
            hoja_modificada_4[datos_parametros_2[key][1] + str(b)] = valor

            b = b + 1
        a = a + 5

    ruta_ultima_4 = ruta_formato_modificado+nombre+".xlsx"            
    libro_modificado_4.save(ruta_ultima_4)
    libro_modificado_4.close()

libro_original.close()    